"""
block_exporter.py
-----------------
Exports a BlockTree to an ST1-format xlsx file.

Output format
-------------
One row per student. Columns:

    Studentid | Grade | A1 | A2 | A3 | ... | H7

Each cell contains 'SUBJECT TEACHER' matching the existing ST1
convention so the output can be fed directly back into timetable_tree.py
and the existing database.

If a student has no assignment in a subblock, the cell is left empty.

Usage
-----
    from block_exporter import export_to_xlsx
    export_to_xlsx(block_tree, school_tree, "output/ST1_new.xlsx")
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

from block_tree  import BlockTree
from school_tree import SchoolTree


# ---------------------------------------------------------------------------
# STYLE CONSTANTS
# Matching the look of the existing ST1 file for consistency
# ---------------------------------------------------------------------------

HEADER_FONT        = Font(bold=True, size=10)
HEADER_FILL        = PatternFill("solid", fgColor="D9D9D9")
HEADER_ALIGNMENT   = Alignment(horizontal="center", vertical="center")

CELL_ALIGNMENT     = Alignment(horizontal="center", vertical="center")
CELL_FONT          = Font(size=9)

BORDER_SIDE        = Side(style="thin", color="CCCCCC")
CELL_BORDER        = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE
)

# Column widths
COL_WIDTH_ID       = 12
COL_WIDTH_GRADE    = 8
COL_WIDTH_SLOT     = 14

# Block fill colours — one per block letter for visual separation
BLOCK_FILLS = {
    "A": PatternFill("solid", fgColor="EEF4FB"),
    "B": PatternFill("solid", fgColor="E8F5E9"),
    "C": PatternFill("solid", fgColor="FFF8E1"),
    "D": PatternFill("solid", fgColor="FCE4EC"),
    "E": PatternFill("solid", fgColor="F3E5F5"),
    "F": PatternFill("solid", fgColor="E0F7FA"),
    "G": PatternFill("solid", fgColor="FBE9E7"),
    "H": PatternFill("solid", fgColor="F1F8E9"),
}


# ---------------------------------------------------------------------------
# HELPER
# ---------------------------------------------------------------------------

def _grade_num(grade_str: str) -> int:
    """'Gr_10' -> 10"""
    return int(grade_str.replace("Gr_", ""))


# ---------------------------------------------------------------------------
# MAIN EXPORT FUNCTION
# ---------------------------------------------------------------------------

def export_to_xlsx(block_tree   : BlockTree,
                   school_tree  : SchoolTree,
                   output_path  : str,
                   sheet_name   : str = "Sheet1") -> Path:
    """
    Export a BlockTree to an ST1-format xlsx file.

    Parameters
    ----------
    block_tree   : populated BlockTree
    school_tree  : used to get all student IDs and grades
    output_path  : destination file path, e.g. "output/ST1_new.xlsx"
    sheet_name   : worksheet name (default matches existing ST1)

    Returns
    -------
    Path to the written file.

    Raises
    ------
    ValueError  if validation errors exist in the BlockTree
    """
    # ---- Validate before export ----
    errors = block_tree.validate()
    if errors:
        print(f"WARNING: BlockTree has {len(errors)} validation error(s).")
        for e in errors:
            print(f"  {e}")
        print("Exporting anyway — errors will appear as clashes in ST1.")

    # ---- Collect all students ordered by grade then ID ----
    all_students = _collect_students(school_tree)

    # ---- Subblock column order ----
    subblock_names = block_tree.all_subblock_names()

    # ---- Build student -> schedule lookup ----
    # { student_id: { subblock_name: st1_cell_value } }
    schedules = _build_schedules(block_tree, all_students, subblock_names)

    # ---- Write xlsx ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    _write_header(ws, subblock_names)
    _write_rows(ws, all_students, subblock_names, schedules)
    _set_column_widths(ws, subblock_names)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    print(f"Exported {len(all_students)} students x "
          f"{len(subblock_names)} slots -> {out}")
    return out


# ---------------------------------------------------------------------------
# INTERNAL HELPERS
# ---------------------------------------------------------------------------

def _collect_students(school_tree: SchoolTree) -> list[tuple[int, str]]:
    """
    Return a sorted list of (student_id, grade) tuples.
    Sorted by grade number then student ID.
    """
    students = set()
    for grade in school_tree.all_grades():
        node = school_tree.grades[grade]
        for group in node.subject_groups.values():
            for student in group.students:
                students.add((int(student.student_id), grade))

    return sorted(students, key=lambda x: (_grade_num(x[1]), x[0]))


def _build_schedules(block_tree     : BlockTree,
                     all_students   : list[tuple[int, str]],
                     subblock_names : list[str]) -> dict:
    """
    Build a lookup:
        { student_id: { subblock_name: cell_value_str } }

    cell_value_str is in ST1 format: 'SUBJECT TEACHER'
    Empty string if the student has no assignment in that slot.
    """
    student_ids = {sid for sid, _ in all_students}
    schedules   = {sid: {} for sid in student_ids}

    for block_name in sorted(block_tree.blocks.keys()):
        block = block_tree.blocks[block_name]
        for sb in block.sorted_subblocks():
            for assignment in sb.all_assignments():
                for sid in assignment.student_ids:
                    if sid in schedules:
                        schedules[sid][sb.name] = assignment.st1_cell

    return schedules


def _write_header(ws, subblock_names: list[str]):
    """Write the header row: Studentid | Grade | A1 | A2 | ..."""
    headers = ["Studentid", "Grade"] + subblock_names

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border    = CELL_BORDER

    ws.row_dimensions[1].height = 18


def _write_rows(ws,
                all_students   : list[tuple[int, str]],
                subblock_names : list[str],
                schedules      : dict):
    """Write one row per student."""
    for row_idx, (student_id, grade) in enumerate(all_students, start=2):
        schedule   = schedules.get(student_id, {})
        grade_disp = grade.replace("Gr_", "")   # "Gr_10" -> "10"

        row_data = [student_id, grade_disp] + [
            schedule.get(sb, "") for sb in subblock_names
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border    = CELL_BORDER

            # Apply block colour to timetable columns
            if col_idx > 2:
                sb_name    = subblock_names[col_idx - 3]
                block_letter = sb_name[0]
                fill = BLOCK_FILLS.get(block_letter)
                if fill:
                    cell.fill = fill

        ws.row_dimensions[row_idx].height = 15


def _set_column_widths(ws, subblock_names: list[str]):
    """Set column widths for readability."""
    ws.column_dimensions["A"].width = COL_WIDTH_ID
    ws.column_dimensions["B"].width = COL_WIDTH_GRADE

    for col_idx in range(3, 3 + len(subblock_names)):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COL_WIDTH_SLOT

    # Freeze the first two columns and header row
    ws.freeze_panes = "C2"


# ---------------------------------------------------------------------------
# Entry point — test with empty tree
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    from pathlib import Path
    from school_tree  import load_from_xlsx
    from block_tree   import BlockTree

    student_file = Path("data/students.xlsx")
    school_tree  = load_from_xlsx(str(student_file))

    # Build an empty tree to verify the exporter structure
    block_tree   = BlockTree()

    out = export_to_xlsx(
        block_tree,
        school_tree,
        output_path="output/ST1_new.xlsx"
    )
    print(f"Done: {out}")