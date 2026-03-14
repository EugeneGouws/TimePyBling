"""
block_exporter.py
-----------------
Exports a BlockTree to an ST1-format xlsx file.

Output format
-------------
One row per student. Columns:

    Studentid | Grade | A1 | A2 | A3 | ... | H7

Each cell contains 'SUBJECT TEACHER' matching the existing ST1
convention so the output can be fed directly back into timetable_tree.py
and the existing database.

If a student has no assignment in a subblock, the cell is left empty.

Usage
-----
    from block_exporter import export_to_xlsx
    export_to_xlsx(block_tree, "output/ST1_new.xlsx")

Note
----
Students and their grades are derived directly from the BlockTree.
No external student list (school_tree) is required.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

from core.block_tree import BlockTree


# ---------------------------------------------------------------------------
# STYLE CONSTANTS
# ---------------------------------------------------------------------------

HEADER_FONT      = Font(bold=True, size=10)
HEADER_FILL      = PatternFill("solid", fgColor="D9D9D9")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

CELL_ALIGNMENT   = Alignment(horizontal="center", vertical="center")
CELL_FONT        = Font(size=9)

BORDER_SIDE      = Side(style="thin", color="CCCCCC")
CELL_BORDER      = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE,  bottom=BORDER_SIDE
)

COL_WIDTH_ID    = 12
COL_WIDTH_GRADE = 8
COL_WIDTH_SLOT  = 14

BLOCK_FILLS = {
    "A": PatternFill("solid", fgColor="EEF4FB"),
    "B": PatternFill("solid", fgColor="E8F5E9"),
    "C": PatternFill("solid", fgColor="FFF8E1"),
    "D": PatternFill("solid", fgColor="FCE4EC"),
    "E": PatternFill("solid", fgColor="F3E5F5"),
    "F": PatternFill("solid", fgColor="E0F7FA"),
    "G": PatternFill("solid", fgColor="FBE9E7"),
    "H": PatternFill("solid", fgColor="F1F8E9"),
}


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def _grade_num(grade) -> int:
    """
    Convert a grade value to an integer for sorting.
    Handles all formats produced by timetable_tree and block_tree:
        "Gr_10"  ->  10
        "10"     ->  10
        "08"     ->   8
        8        ->   8
    """
    return int(str(grade).replace("Gr_", ""))


def _grade_display(grade) -> str:
    """
    Convert a grade value to a plain number string for the xlsx cell.
        "Gr_10"  ->  "10"
        "08"     ->  "08"
        8        ->  "8"
    """
    return str(grade).replace("Gr_", "")


# ---------------------------------------------------------------------------
# MAIN EXPORT FUNCTION
# ---------------------------------------------------------------------------

def export_to_xlsx(block_tree : BlockTree,
                   output_path: str,
                   sheet_name : str = "Sheet1") -> Path:
    """
    Export a BlockTree to an ST1-format xlsx file.

    Students and their grades are read directly from the SlotAssignments
    inside the tree — no external student list is needed.

    Parameters
    ----------
    block_tree   : populated BlockTree (from timetable_tree_to_block_tree
                   or from block_builder)
    output_path  : destination file path, e.g. "output/ST1_new.xlsx"
    sheet_name   : worksheet name (default "Sheet1" matches existing ST1)

    Returns
    -------
    Path to the written file.
    """
    errors = block_tree.validate()
    if errors:
        print(f"WARNING: BlockTree has {len(errors)} validation error(s).")
        for e in errors:
            print(f"  {e}")
        print("Exporting anyway — errors will appear as clashes in ST1.")

    all_students   = _collect_students(block_tree)
    subblock_names = block_tree.all_subblock_names()
    schedules      = _build_schedules(block_tree, all_students, subblock_names)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    _write_header(ws, subblock_names)
    _write_rows(ws, all_students, subblock_names, schedules)
    _set_column_widths(ws, subblock_names)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    print(f"Exported {len(all_students)} students x "
          f"{len(subblock_names)} slots -> {out}")
    return out


# ---------------------------------------------------------------------------
# INTERNAL HELPERS
# ---------------------------------------------------------------------------

def _collect_students(block_tree: BlockTree) -> list[tuple[int, str]]:
    """
    Build a sorted list of (student_id, grade) from the BlockTree itself.

    A student appears in many SlotAssignments (one per subblock) but always
    with the same grade, so we keep the first occurrence and move on.

    Sorted by grade number then student ID, matching the original ST1 layout.
    """
    students: dict[int, str] = {}

    for _sb_name, assignment in block_tree.all_assignments():
        for sid in assignment.student_ids:
            if sid not in students:
                students[sid] = assignment.grade

    return sorted(
        students.items(),
        key=lambda x: (_grade_num(x[1]), x[0])
    )


def _build_schedules(block_tree    : BlockTree,
                     all_students  : list[tuple[int, str]],
                     subblock_names: list[str]) -> dict:
    """
    Build { student_id: { subblock_name: 'SUBJECT TEACHER' } }.
    Empty string where the student has no assignment in a slot.
    """
    student_ids = {sid for sid, _ in all_students}
    schedules   = {sid: {} for sid in student_ids}

    for block_name in sorted(block_tree.blocks.keys()):
        block = block_tree.blocks[block_name]
        for sb in block.sorted_subblocks():
            for assignment in sb.all_assignments():
                for sid in assignment.student_ids:
                    if sid in schedules:
                        schedules[sid][sb.name] = assignment.st1_cell

    return schedules


def _write_header(ws, subblock_names: list[str]):
    headers = ["Studentid", "Grade"] + subblock_names

    for col_idx, header in enumerate(headers, start=1):
        cell           = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border    = CELL_BORDER

    ws.row_dimensions[1].height = 18


def _write_rows(ws,
                all_students  : list[tuple[int, str]],
                subblock_names: list[str],
                schedules     : dict):
    for row_idx, (student_id, grade) in enumerate(all_students, start=2):
        schedule    = schedules.get(student_id, {})
        grade_disp  = _grade_display(grade)

        row_data = [student_id, grade_disp] + [
            schedule.get(sb, "") for sb in subblock_names
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = CELL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border    = CELL_BORDER

            if col_idx > 2:
                block_letter = subblock_names[col_idx - 3][0]
                fill = BLOCK_FILLS.get(block_letter)
                if fill:
                    cell.fill = fill

        ws.row_dimensions[row_idx].height = 15


def _set_column_widths(ws, subblock_names: list[str]):
    ws.column_dimensions["A"].width = COL_WIDTH_ID
    ws.column_dimensions["B"].width = COL_WIDTH_GRADE

    for col_idx in range(3, 3 + len(subblock_names)):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = COL_WIDTH_SLOT

    ws.freeze_panes = "C2"


# ---------------------------------------------------------------------------
# Standalone test
# ---------------------------------------------------------------------------

if __name__ == "__main__":

    st_file = Path("../data/ST1.xlsx")
    if not st_file.exists():
        st_file = Path("ST1.xlsx")

    if not st_file.exists():
        print("No ST1.xlsx found.")
        raise SystemExit(1)

    print(f"Loading: {st_file}")
    # Once timetable_tree_to_block_tree() exists, use it here:
    # timetable_tree = build_timetable_tree_from_file(st_file)
    # block_tree = timetable_tree_to_block_tree(timetable_tree)
    # export_to_xlsx(block_tree, "output/ST1_new.xlsx")
    print("Ready — wire up timetable_tree_to_block_tree() to run a full export.")